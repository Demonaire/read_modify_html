'''
Note:
1. There should be a data file to be edited with name: data.xlsx
2. There should be a file name: textToBePut.txt in the same folder
3. Installed BS4 library
4. Installed openpyxl module
5. The new ammended data is saved as dataWithAddition.xlsx
6. NOt recommended to edit the code unless you're confident enough

'''

from openpyxl import load_workbook
from bs4 import BeautifulSoup
import openpyxl
import re
#import column AL to python
workbook = load_workbook(filename='data.xlsx',data_only=True)
sheet = workbook.active
no_fill = openpyxl.styles.PatternFill(fill_type=None)
lst = ['A','B','C','D','F','I','J','K','L','M','N','O','P','Q','R','S','T','U','V','W','X','Y','Z','AA','AB','AC','AD','AE','AG','AH','AI','AJ','AL','AS','AT','AU','AV','AW','AX','AY','AZ','BA','BB','BC','BD','BE','BF','BG','BH','BI','BJ','BK']
# import the text to be inserted
textToBePut = ""
with open('textToBePut.txt') as file:
    textToBePut = file.read()
textToBePut = "".join(line.strip() for line in textToBePut.split("\n")) #remove whitespaces
for i, cell in enumerate(sheet['AL']):
    if i== 0:
        continue
    if cell.value == None:
        
        ##removing blanks filled cells to match original file format
        for row in lst:
            sheet[row+str(i+1)]=""
            sheet[row+str(i+1)].fill=no_fill     
        ############################################################
        
        continue
        
    ##removing blanks filled cells to match original file format    
    if i==12 or (i-12)%32==0:
        for row in lst:
            sheet[row+str(i+1)]=""
            sheet[row+str(i+1)].fill=no_fill
    if i==1291 or (i-1291)%32==0:
        for row in lst:
            sheet[row+str(i+1)]=""
            sheet[row+str(i+1)].fill=no_fill
    ############################################################
    soup = BeautifulSoup(cell.value, "html.parser")
    soup_chart = BeautifulSoup(textToBePut, "html.parser")
    div = soup.find_all('strong', string="About Us")
    if (len(div) > 0 ): 
        target_tag = div[0].parent.parent
    target_tag.insert_before(soup_chart)
    pattern='<p align="justify">'+ " Clara Pucci's Simulated Diamonds"
    replacement='<p align="center">'+ " Clara Pucci's Simulated Diamonds"
    soup=re.sub(pattern,replacement,str(soup))
    sheet['AL'+str(i+1)] = str(soup)

workbook.save(filename="dataWithAddition.xlsx")